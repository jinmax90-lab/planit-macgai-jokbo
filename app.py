"""
플래닛학원 학생관리 시스템 (맥가이 버전)
Streamlit 웹앱
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 한국 시간대 (UTC+9)
KST = timezone(timedelta(hours=9))

# ============================================================
# 설정
# ============================================================
MAX_GROUP_SIZE = 30

ALL_COLUMNS = [
    '학교', '학년', '이름', '학생번호', '선생님', '과목',
    '배포그룹', '그룹순번', '족보ID',
    '상태', '학생HP', '학부모HP', '메모'
]

# 스타일
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ============================================================
# 유틸리티 함수
# ============================================================
def safe_str(value):
    """NaN 안전 문자열 변환"""
    if pd.isna(value):
        return ''
    return str(value).strip()

def safe_student_id(value):
    """학생번호 안전 변환 (5자리 유지)"""
    if pd.isna(value):
        return ''
    if isinstance(value, float):
        value = int(value)
    return str(value).strip().zfill(5)

def clean_phone(value):
    """전화번호 정리"""
    if pd.isna(value):
        return ''
    return str(value).strip()

def make_key(student_id, teacher, subject):
    """강좌 키: 학생번호+선생님+과목"""
    return f"{safe_student_id(student_id)}_{safe_str(teacher)}_{safe_str(subject)}"

def make_match_key(name, parent_phone):
    """매칭 키: 이름+학부모전화번호"""
    return f"{safe_str(name)}_{clean_phone(parent_phone)}"

def parse_class_name(class_str):
    """
    맥가이 학급명 파싱: '고2 물리A 수8 송승용T' → ('고2 물리', '송승용')
    """
    if not class_str or pd.isna(class_str):
        return None, None
    
    parts = str(class_str).strip().split()
    if len(parts) < 4:
        return None, None
    
    # 마지막이 선생님 (xxxT)
    teacher_part = parts[-1]
    if teacher_part.endswith('T'):
        teacher = teacher_part[:-1]
    else:
        teacher = teacher_part
    
    # 첫 두 부분이 학년+과목 (고2 물리)
    grade = parts[0]  # 고2
    subject_with_class = parts[1]  # 물리A
    
    # 과목에서 반 제거 (물리A → 물리)
    subject = ''.join([c for c in subject_with_class if not c.isalpha() or c in '가나다라마바사아자차카타파하'])
    if subject_with_class:
        # 마지막 알파벳이 반이면 제거
        if subject_with_class[-1].isalpha() and subject_with_class[-1].isupper():
            subject = subject_with_class[:-1]
        else:
            subject = subject_with_class
    
    full_subject = f"{grade} {subject}"
    
    return full_subject, teacher

def parse_all_classes(class_str):
    """
    여러 과목 파싱: '고2 물리A 수8 송승용T／고2 화학D 금8 송승용T'
    → [('고2 물리', '송승용'), ('고2 화학', '송승용')]
    """
    if not class_str or pd.isna(class_str):
        return []
    
    results = []
    classes = str(class_str).split('／')
    
    for cls in classes:
        cls = cls.strip()
        if cls:
            subject, teacher = parse_class_name(cls)
            if subject and teacher:
                results.append((subject, teacher))
    
    return results

# ============================================================
# 엑셀 저장 함수
# ============================================================
def save_to_excel(df, color_info=None):
    """DataFrame을 엑셀로 저장하고 BytesIO 반환"""
    wb = Workbook()
    ws = wb.active
    ws.title = "학생관리"
    
    if color_info is None:
        color_info = {}
    
    all_cols = list(df.columns)
    
    # 헤더
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    
    col_to_idx = {col: idx + 1 for idx, col in enumerate(all_cols)}
    jokbo_col = col_to_idx.get('족보ID', None)
    student_id_col = col_to_idx.get('학생번호', None)
    
    # 데이터
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            if pd.isna(value):
                value = ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # 학생번호 5자리 텍스트
        if student_id_col:
            cell = ws.cell(row=row_idx, column=student_id_col)
            cell.number_format = '@'
            if cell.value not in ['', None]:
                val = cell.value
                if isinstance(val, float):
                    val = int(val)
                cell.value = str(val).zfill(5)
        
        # 족보ID 텍스트
        if jokbo_col:
            cell = ws.cell(row=row_idx, column=jokbo_col)
            cell.number_format = '@'
            if cell.value not in ['', None]:
                cell.value = str(cell.value)
        
        # 색상
        df_idx = row_idx - 2
        if df_idx in color_info:
            colors = color_info[df_idx]
            if colors.get('이름') == 'green':
                ws.cell(row=row_idx, column=col_to_idx['이름']).fill = GREEN_FILL
            elif colors.get('이름') == 'red':
                ws.cell(row=row_idx, column=col_to_idx['이름']).fill = RED_FILL
            elif colors.get('이름') == 'yellow':
                ws.cell(row=row_idx, column=col_to_idx['이름']).fill = YELLOW_FILL
            if colors.get('족보ID') == 'purple':
                ws.cell(row=row_idx, column=col_to_idx['족보ID']).fill = PURPLE_FILL
    
    # 컬럼 너비
    col_widths = {
        '학교': 10, '학년': 6, '이름': 12, '학생번호': 8, '선생님': 8, '과목': 12,
        '배포그룹': 10, '그룹순번': 10, '족보ID': 12,
        '상태': 10, '학생HP': 16, '학부모HP': 16, '메모': 15
    }
    for col_idx, col_name in enumerate(all_cols, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 12)
    
    # 필터, 틀고정
    last_col = get_column_letter(len(all_cols))
    ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"
    ws.freeze_panes = 'A2'
    
    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# 메인 처리 함수
# ============================================================
def process_update(df_macguy, df_master=None):
    """맥가이 회원명단 + 기존 마스터 → 새 마스터"""
    
    logs = []
    
    # 1. 맥가이 데이터 파싱
    logs.append("📥 맥가이 회원명단 처리 중...")
    
    new_records = []
    for _, row in df_macguy.iterrows():
        name = safe_str(row.get('학생명', ''))
        student_id = safe_student_id(row.get('학생번호', ''))
        school = safe_str(row.get('학교', ''))
        grade = safe_str(row.get('학년', ''))
        student_hp = clean_phone(row.get('학생연락처', ''))
        parent_hp = clean_phone(row.get('학부모연락처', ''))
        class_str = safe_str(row.get('학급', ''))
        status_raw = safe_str(row.get('상태', ''))
        
        # 상태 변환
        if status_raw in ['재원', '신입']:
            status = '수강중'
        elif status_raw == '휴원':
            status = '휴원'
        elif status_raw == '퇴원':
            status = '퇴원'
        else:
            status = '수강중'
        
        # 학급 파싱 (여러 과목)
        classes = parse_all_classes(class_str)
        
        if classes:
            for subject, teacher in classes:
                new_records.append({
                    '학교': school,
                    '학년': grade,
                    '이름': name,
                    '학생번호': student_id,
                    '선생님': teacher,
                    '과목': subject,
                    '배포그룹': '',
                    '그룹순번': '',
                    '족보ID': '',
                    '상태': status,
                    '학생HP': student_hp,
                    '학부모HP': parent_hp,
                    '메모': '',
                    '_match_key': make_match_key(name, parent_hp)
                })
        else:
            # 학급이 없는 경우 (과목미정)
            new_records.append({
                '학교': school,
                '학년': grade,
                '이름': name,
                '학생번호': student_id,
                '선생님': '',
                '과목': '',
                '배포그룹': '',
                '그룹순번': '',
                '족보ID': '',
                '상태': '과목미정' if status == '수강중' else status,
                '학생HP': student_hp,
                '학부모HP': parent_hp,
                '메모': '',
                '_match_key': make_match_key(name, parent_hp)
            })
    
    df_new = pd.DataFrame(new_records)
    # 모든 컬럼을 object 타입으로 변환 (pandas Arrow 호환성)
    for col in df_new.columns:
        df_new[col] = df_new[col].astype(object)
    logs.append(f"  → 맥가이 레코드: {len(df_new)}건")
    
    # 2. 기존 마스터에서 족보ID, 배포그룹 복원
    new_course_keys = set()
    color_info = {}
    
    if df_master is not None and len(df_master) > 0:
        logs.append("📂 기존 마스터에서 데이터 복원 중...")
        
        # 기존 마스터의 매칭키 → 데이터 맵
        master_data = {}
        for _, row in df_master.iterrows():
            name = safe_str(row.get('이름', ''))
            parent_hp = clean_phone(row.get('학부모HP', ''))
            teacher = safe_str(row.get('선생님', ''))
            subject = safe_str(row.get('과목', ''))
            
            match_key = make_match_key(name, parent_hp)
            course_key = f"{match_key}_{teacher}_{subject}"
            
            master_data[course_key] = {
                '족보ID': safe_str(row.get('족보ID', '')),
                '배포그룹': row.get('배포그룹', ''),
                '그룹순번': row.get('그룹순번', ''),
                '메모': safe_str(row.get('메모', ''))
            }
        
        # 새 데이터에 복원
        restored_count = 0
        for idx, row in df_new.iterrows():
            match_key = row['_match_key']
            teacher = safe_str(row.get('선생님', ''))
            subject = safe_str(row.get('과목', ''))
            course_key = f"{match_key}_{teacher}_{subject}"
            
            if course_key in master_data:
                old_data = master_data[course_key]
                if old_data['족보ID']:
                    df_new.at[idx, '족보ID'] = str(old_data['족보ID'])
                if old_data['배포그룹'] not in ['', None] and pd.notna(old_data['배포그룹']):
                    df_new.at[idx, '배포그룹'] = str(int(float(old_data['배포그룹'])))
                if old_data['메모']:
                    df_new.at[idx, '메모'] = str(old_data['메모'])
                restored_count += 1
            else:
                # 기존에 없던 신규 강좌
                if teacher and subject:
                    new_course_keys.add(idx)
        
        logs.append(f"  → 기존 데이터 복원: {restored_count}건")
        logs.append(f"  → 신규 강좌: {len(new_course_keys)}건")
    else:
        logs.append("📂 기존 마스터 없음 - 초기 생성 모드")
        for idx, row in df_new.iterrows():
            if row.get('선생님') and row.get('과목'):
                new_course_keys.add(idx)
    
    # 3. 배포그룹 배정 (없는 것만)
    logs.append("📊 배포그룹 배정 중...")
    
    from collections import defaultdict
    group_counts = defaultdict(lambda: defaultdict(int))
    
    # 기존 배포그룹 카운트
    for idx, row in df_new.iterrows():
        teacher = safe_str(row.get('선생님', ''))
        subject = safe_str(row.get('과목', ''))
        bg = row.get('배포그룹', '')
        status = safe_str(row.get('상태', ''))
        
        if teacher and subject and bg not in ['', None] and pd.notna(bg) and status not in ['퇴원', '과목미정', '휴원']:
            group_key = f"{teacher}_{subject}"
            try:
                group_counts[group_key][int(float(bg))] += 1
            except:
                pass
    
    # 신규 배정
    assigned_count = 0
    for idx, row in df_new.iterrows():
        bg = row.get('배포그룹', '')
        status = safe_str(row.get('상태', ''))
        teacher = safe_str(row.get('선생님', ''))
        subject = safe_str(row.get('과목', ''))
        
        if (pd.notna(bg) and bg != '') or status in ['퇴원', '과목미정', '휴원'] or not teacher or not subject:
            continue
        
        group_key = f"{teacher}_{subject}"
        for group_num in range(1, 100):
            if group_counts[group_key][group_num] < MAX_GROUP_SIZE:
                df_new.at[idx, '배포그룹'] = str(group_num)
                group_counts[group_key][group_num] += 1
                assigned_count += 1
                break
    
    logs.append(f"  → 신규 배정: {assigned_count}건")
    
    # 4. 그룹순번 계산
    df_new['그룹순번'] = ''
    valid = df_new[
        (df_new['선생님'].notna()) & (df_new['선생님'] != '') &
        (df_new['과목'].notna()) & (df_new['과목'] != '') &
        (df_new['배포그룹'].notna()) & (df_new['배포그룹'] != '') &
        (~df_new['상태'].isin(['퇴원', '휴원']))
    ]
    
    if len(valid) > 0:
        grouped = valid.groupby(['선생님', '과목', '배포그룹'])
        for (teacher, subject, group), grp in grouped:
            sorted_grp = grp.sort_values('이름')
            for order, (grp_idx, _) in enumerate(sorted_grp.iterrows(), 1):
                df_new.at[grp_idx, '그룹순번'] = str(order)
    
    # 5. 정렬
    df_new = df_new.sort_values(['이름', '선생님', '과목']).reset_index(drop=True)
    
    # 6. 색상 정보
    for idx, row in df_new.iterrows():
        row_colors = {}
        status = safe_str(row.get('상태', ''))
        jokbo_id = safe_str(row.get('족보ID', ''))
        teacher = safe_str(row.get('선생님', ''))
        
        # 신규 강좌 (초록)
        if idx in new_course_keys:
            row_colors['이름'] = 'green'
        
        # 퇴원/휴원 (빨강)
        if status in ['퇴원', '휴원']:
            row_colors['이름'] = 'red'
        
        # 과목미정 (노랑)
        if status == '과목미정':
            row_colors['이름'] = 'yellow'
        
        # 족보ID 없음 (보라)
        if not jokbo_id and status == '수강중' and teacher:
            row_colors['족보ID'] = 'purple'
        
        if row_colors:
            color_info[idx] = row_colors
    
    # 7. _match_key 컬럼 제거, 컬럼 정리
    if '_match_key' in df_new.columns:
        df_new = df_new.drop(columns=['_match_key'])
    
    # 컬럼 순서 정리
    for col in ALL_COLUMNS:
        if col not in df_new.columns:
            df_new[col] = ''
    df_new = df_new[ALL_COLUMNS]
    
    # 8. 통계
    total_active = len(df_new[df_new['상태'] == '수강중'])
    total_retired = len(df_new[df_new['상태'].isin(['퇴원', '휴원'])])
    total_pending = len(df_new[df_new['상태'] == '과목미정'])
    
    logs.append("")
    logs.append("✅ 처리 완료!")
    logs.append(f"  - 수강중: {total_active}건")
    logs.append(f"  - 퇴원/휴원: {total_retired}건")
    logs.append(f"  - 과목미정: {total_pending}건")
    
    return df_new, color_info, logs

# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(
    page_title="플래닛학원 족보ID관리",
    page_icon="🏫",
    layout="centered"
)

# 버전 관리 (화면에 표시 안함)
# v1.2

# 업데이트 시점 표시
st.markdown('<p style="color: #666; font-size: 12px; text-align: right; margin-bottom: 0;">2026-04-03 21:00 업데이트</p>', unsafe_allow_html=True)

st.markdown("#### 🏫 플래닛학원 족보ID관리")
st.caption("맥가이 - 회원명단 엑셀파일 → 학생관리 최종파일")

st.divider()

# 파일 업로드
col1, col2 = st.columns(2)

with col1:
    st.markdown("##### 📁 맥가이 - 회원명단")
    uploaded_macguy = st.file_uploader(
        "회원명단 파일 업로드",
        type=['xls', 'xlsx'],
        key='macguy',
        help="맥가이에서 다운받은 회원명단 파일"
    )

with col2:
    st.markdown("##### 📁 기존 학생관리 파일")
    uploaded_master = st.file_uploader(
        "기존 최종파일 업로드 (선택)",
        type=['xlsx'],
        key='master',
        help="족보ID, 배포그룹이 있는 기존 파일 (없으면 새로 생성)"
    )

st.divider()

# 실행 버튼
if st.button("🚀 업데이트 실행", type="primary", use_container_width=True):
    if not uploaded_macguy:
        st.error("❌ 맥가이 - 회원명단 엑셀파일을 업로드해주세요.")
    else:
        with st.spinner("처리 중..."):
            try:
                # 맥가이 파일 읽기
                df_macguy = pd.read_html(uploaded_macguy)[0]
                
                # 기존 마스터 읽기
                df_master = None
                if uploaded_master:
                    df_master = pd.read_excel(uploaded_master)
                
                # 처리
                df_result, color_info, logs = process_update(df_macguy, df_master)
                
                # 로그 출력
                for log in logs:
                    if log.startswith("✅"):
                        st.success(log)
                    elif log.startswith("❌"):
                        st.error(log)
                    elif log.startswith("  -"):
                        st.write(log)
                    elif log:
                        st.info(log)
                
                # 결과 저장 (한국 시간 기준)
                timestamp = datetime.now(KST).strftime("%y%m%d_%H%M")
                filename = f"{timestamp}_족보ID 관리.xlsx"
                
                excel_data = save_to_excel(df_result, color_info)
                
                st.divider()
                st.subheader("📥 결과 다운로드")
                
                st.download_button(
                    label=f"📥 {filename} 다운로드",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
                
                # 미리보기
                with st.expander("📋 결과 미리보기"):
                    st.dataframe(df_result.head(20))
                    
            except Exception as e:
                st.error(f"❌ 오류 발생: {str(e)}")
                import traceback
                st.code(traceback.format_exc())

# 사용 안내
with st.expander("📖 사용 안내"):
    st.markdown("""
    ### 사용 방법
    1. **맥가이에서 회원명단** 다운로드
    2. **기존 학생관리 엑셀파일** 업로드 (족보ID, 배포그룹 보존용)
    3. **업데이트 실행** 버튼 클릭
    4. **결과 다운로드**
    
    ### 색상 안내
    - 🟢 **초록**: 신규 강좌
    - 🔴 **빨강**: 퇴원/휴원
    - 🟡 **노랑**: 과목미정
    - 🟣 **보라**: 족보ID 미입력
    
    ### 주의사항
    - 기존 파일을 올리면 **족보ID, 배포그룹**이 자동으로 복원됩니다
    - 이름+학부모전화번호로 학생을 매칭합니다
    """)
